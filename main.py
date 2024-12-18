import re
import copy
import tkinter as tk
import classes as cl
import openpyxl as op
from datetime import datetime
from tkinter import filedialog as fd
from openpyxl.styles import PatternFill

list_bekasovo = ('ERD-577', 'ERD-578', 'ERD-579', 'ERD-580', 'ERD-582',
                 'ERD-583', 'ERD-584', 'ERD-585', 'ERD-587', 'ERD-588')

list_ng4_14 = ('ERD-444', 'ERD-445', 'ERD-446', 'ERD-447', 'ERD-448', 'ERD-449', 'ERD-450', 'ERD-451', 'ERD-452',
               'ERD-453', 'ERD-454', 'ERD-455', 'ERD-456', 'ERD-457', 'ERD-458', 'ERD-569', 'ERD-460')

filename = ''
action_list = ['Обычная', 'Выгрузка по НГЧ', 'Выгрузка по температурам', 'Выгрузка Бексаово, НГЧ14']
list_info_for_work = tuple()
answer_file_name = ""


def open_file_name():
    global filename
    filename = str(fd.askopenfilenames())
    filename = filename.replace('/', '//').replace('(', '').replace(')', '').replace("'", '').replace(",", '')
    label_select_file.set_text('label_select_file', filename)
    return filename


def clear_window(list_widgets):
    dict_label = cl.MyLabel.label_list
    dict_entry = cl.MyEntry.entry_list
    dict_checkbutton = cl.Checkbuttons.checkbutton_list

    for widget_i in list_widgets:
        temp_entry = dict_entry.get(widget_i)
        temp_label = dict_label.get(widget_i)
        temp_checkbutton = dict_checkbutton.get(widget_i)
        if temp_entry is not None:
            temp_entry.place_forget()

        if temp_label is not None:
            temp_label.place_forget()

        if temp_checkbutton is not None:
            temp_checkbutton[0].place_forget()


def by_normal(file):
    book = op.load_workbook(file)
    sheet = book.active
    max_vols = sheet.max_row
    temp_list = book.create_sheet('Norml')

    for row_i in range(1, max_vols + 1):
        str_row = str(sheet.cell(row=row_i, column=2).value)
        if '/00' not in str_row:
            _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=row_i, max_row=row_i, values_only=True)]

    book.remove(sheet)
    temp_list.column_dimensions['A'].width = 18
    temp_list.column_dimensions['C'].width = 50
    now_time = datetime.now()
    file_name = f'Общая {now_time.day}.{now_time.month}.{now_time.year}.xlsx'
    file_path = fd.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel file", '*.xlsx')], initialfile=file_name)
    book.save(file_path)
    file_name += ' готова'
    return file_name


def by_temperature(file, min_pom, max_pom, min_cont, max_cont):
    book = op.load_workbook(file)
    sheet = book.active
    max_vols = sheet.max_row
    fill_blue = PatternFill('solid', fgColor='93ccdd')
    fill_red = PatternFill('solid', fgColor='e6b8b8')
    fill_purple = PatternFill('solid', fgColor='cdc0da')

    temp_list = book.create_sheet('Norml')
    for row_i in range(1, max_vols + 1):
        str_row = str(sheet.cell(row=row_i, column=2).value)
        if '/00' not in str_row:
            _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=row_i, max_row=row_i, values_only=True)]

    book.remove(sheet)
    sheet = book.active
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['C'].width = 50
    for i_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=i_id, column=13).value)
            if num == 'SNMP No-Such-Instance':
                continue
            elif num <= int(min_pom):
                sheet[f'M{i_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'M{i_id}'].fill = fill_red
        except:
            continue

    for j_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=j_id, column=14).value)
            if num == 'SNMP No-Such-Instance':
                continue

            elif num <= int(min_pom):
                sheet[f'N{j_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'N{j_id}'].fill = fill_red
        except:
            continue

    for e_id in range(max_vols + 1):
        try:
            num = str(sheet.cell(row=e_id, column=7).value)
            if num == 'Есть':
                sheet[f'G{e_id}'].fill = fill_purple
        except:
            continue

    for f_id in range(max_vols + 1):
        try:

            num = int(sheet.cell(row=f_id, column=6).value)
            if num == 'SNMP No-Such-Instance':
                continue
            elif num <= (min_cont):
                sheet[f'F{f_id}'].fill = fill_blue
            elif num >= (max_cont):
                sheet[f'F{f_id}'].fill = fill_red
        except:
            continue

    now_time = datetime.now()
    file_name = f'Выгрузка по температурам {now_time.day}.{now_time.month}.{now_time.year}.xlsx'
    file_path = fd.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel file", '*.xlsx')], initialfile=file_name)
    book.save(file_path)

    file_name += ' готова'
    return file_name


def by_ng4(file):
    checkbutton_list = cl.Checkbuttons.checkbutton_list
    separate = checkbutton_list['По отдельным файлам (Выгрузки для отправки на почты)'][1].get()

    temp_path = fd.asksaveasfilename(defaultextension='.xlsx',
                                     filetypes=[("Excel file", '*.xlsx')],
                                     initialfile='Выберать эта папку (Наименование файла указывать не нужно)'
                                     )
    if separate:




        for keys, value in checkbutton_list.items():



            if int(value[1].get()) and keys != 'По отдельным файлам (Выгрузки для отправки на почты)':
                copy_path = copy.copy(temp_path)
                book = op.load_workbook(file)
                sheet = book.active
                max_vols = sheet.max_row
                file_name = keys
                print(file_name)
                temp_list = book.create_sheet(keys)
                temp_list.column_dimensions['A'].width = 18
                temp_list.column_dimensions['C'].width = 50
                _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]

                for i_row in range(2, max_vols):
                    if str(sheet.cell(row=i_row, column=2).value) == keys:
                        _ = [temp_list.append(row) for row in
                             sheet.iter_rows(min_row=i_row, max_row=i_row, values_only=True)]

                if keys == 'НГЧ-5':
                    _ = [temp_list.append(row) for row in
                         sheet.iter_rows(min_row=max_vols, max_row=max_vols, values_only=True)]
                    # В исходной выгрузке косяк с наименованием НГЧ,
                    # по этому добавляется ещё самая последняя строка(с косячным наименованием)

                book.remove(sheet)

                file_name = f'{file_name}.xlsx'
                print(temp_path)
                copy_path = re.sub(r'/+','//', copy_path)
                copy_path = copy_path.split('//')
                file_path = ''
                for i in range(len(copy_path)-1):
                    file_path += copy_path[i] + '//'

                file_path = file_path + file_name
                print(file_path)
                book.save(file_path)
                book.close()

        return "Выгрузки готовы"
    else:
        copy_path = copy.copy(temp_path)
        book = op.load_workbook(file)
        sheet = book.active
        max_vols = sheet.max_row
        file_name = ''
        for keys, value in checkbutton_list.items():

            if int(value[1].get()) and keys != 'По отдельным файлам (Выгрузки для отправки на почты)':
                file_name += f'{keys}, '
                temp_list = book.create_sheet(keys)
                temp_list.column_dimensions['A'].width = 18
                temp_list.column_dimensions['C'].width = 50
                _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
                for i_row in range(2, max_vols):
                    if str(sheet.cell(row=i_row, column=2).value) == keys:
                        _ = [temp_list.append(row) for row in
                             sheet.iter_rows(min_row=i_row, max_row=i_row, values_only=True)]

                if keys == 'НГЧ-5':
                    _ = [temp_list.append(row) for row in
                         sheet.iter_rows(min_row=max_vols, max_row=max_vols, values_only=True)]
                    # В исходной выгрузке косяк с наименованием НГЧ,
                    # по этому добавляется ещё самая последняя строка(с косячным наименованием)

        book.remove(sheet)
        file_name = f'{file_name}.xlsx'
        copy_path = re.sub(r'/+', '//', copy_path)
        copy_path = copy_path.split('//')
        file_path = ''
        for i in range(len(copy_path) - 1):
            file_path += copy_path[i] + '//'

        file_path = file_path + file_name
        print(file_path)
        book.save(file_path)
        file_name += ' готова'
        return file_name



def by_bek_and_ng4_14(file):
    global list_bekasovo
    global list_ng4_14
    book = op.load_workbook(file)
    sheet = book.active
    max_vols = sheet.max_row

    temp_list = book.create_sheet("Стр 1")
    _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
    temp_list.column_dimensions['A'].width = 18
    temp_list.column_dimensions['C'].width = 50

    for id_i in range(2, max_vols):
        erd_now = str(sheet.cell(row=id_i, column=5).value)
        if erd_now in list_bekasovo:
            _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=id_i, max_row=id_i, values_only=True)]

    _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]

    for id_i in range(2, max_vols):
        erd_now = str(sheet.cell(row=id_i, column=5).value)
        if erd_now in list_ng4_14:
            _ = [temp_list.append(row) for row in sheet.iter_rows(min_row=id_i, max_row=id_i, values_only=True)]

    book.remove(sheet)

    file_name = f'Выгрузка Бекасово и НГЧ14.xlsx'
    file_path = fd.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel file", '*.xlsx')], initialfile=file_name)
    book.save(file_path)
    file_name += ' готова'
    return file_name


def interface_for_temperature(root):
    global list_info_for_work
    clear_window(list_info_for_work)

    settings = cl.Settings('settings.ini')
    cl.Settings.setting_obj.append(settings)
    # print(settings.settings)

    cl.MyLabel(root, 'label_manual_temp',
               'Выделит цветом критические температуры и др. параметры, уберёт /00, раздвинет столбцы', 25, 125)
    # Лабель и ввод минимальных помещение
    cl.MyLabel(root, 'label_get_min_tem_pom', 'Min t помещения: ', 25, 160)
    cl.MyEntry(root, 'input_min_temp_pom', settings.settings['Temperature']['min_t_pom'], 140, 160)
    # Лабель и ввод максимальных помещение
    cl.MyLabel(root, 'label_get_max_tem_pom', 'Max t помещения: ', 25, 180)
    cl.MyEntry(root, 'input_max_temp_pom', settings.settings['Temperature']['max_t_pom'], 140, 180)

    # Лабель и ввод минимальных контуров
    cl.MyLabel(root, 'label_get_min_temp_cont', 'Min t контура: ', 250, 160)
    cl.MyEntry(root, 'input_min_temp_cont', settings.settings['Temperature']['min_t_cont'], 340, 160)

    # Лабель и ввод максимальных контуров
    cl.MyLabel(root, 'label_get_max_temp_cont', 'Max t контура: ', 250, 180)
    cl.MyEntry(root, 'input_max_temp_cont', settings.settings['Temperature']['max_t_cont'], 340, 180)

    list_info_for_work = (
        'label_manual_temp', 'input_min_temp_pom', 'input_max_temp_pom', 'input_min_temp_cont', 'input_max_temp_cont',
        'label_get_min_tem_pom', 'label_get_max_tem_pom', 'label_get_min_temp_cont',
        'label_get_max_temp_cont')

    return list_info_for_work


def interface_for_NG4(root):
    global list_info_for_work
    clear_window(list_info_for_work)

    cl.MyLabel(root, 'label_manual_temp', 'Вытащит выбранные НГЧ, по отдельным листам или файлам, уберёт /00, раздвинет столбцы',
               25, 125)

    value_ng4_separate = tk.IntVar(root)
    cl.Checkbuttons(root, 'По отдельным файлам (Выгрузки для отправки на почты)', value_ng4_separate, 250, 120 + 30)

    value_ng4_11 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-11', value_ng4_11, 25, 120 + 30)
    value_ng4_13 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-13', value_ng4_13, 25, 140 + 30)
    value_ng4_14 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-14', value_ng4_14, 25, 160 + 30)
    value_ng4_17 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-17', value_ng4_17, 25, 180 + 30)
    value_ng4_19 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-19', value_ng4_19, 130, 120 + 30)
    value_ng4_20 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-20', value_ng4_20, 130, 140 + 30)
    value_ng4_5 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-5', value_ng4_5, 130, 160 + 30)
    value_ng4_6 = tk.IntVar(root)
    cl.Checkbuttons(root, 'НГЧ-6', value_ng4_6, 130, 180 + 30)
    list_info_for_work = (
        'По отдельным файлам (Выгрузки для отправки на почты)', 'label_manual_temp', 'НГЧ-11', 'НГЧ-13', 'НГЧ-14',
        'НГЧ-17',
        'НГЧ-19', 'НГЧ-20', 'НГЧ-5', 'НГЧ-6')

    return list_info_for_work


def interface_for_normal(root):
    global list_info_for_work
    clear_window(list_info_for_work)

    cl.MyLabel(root, 'label_normal_screen',
               'При данном варианте обработки, \nбудут удалены все лишние котлы с "/00", раздвинуты столбцы.',
               30, 130)
    list_info_for_work = ('label_normal_screen',)


def interface_for_becasovo(root):
    global list_info_for_work
    clear_window(list_info_for_work)

    cl.MyLabel(root, 'label_bekasovo',
               'При данном варианте обработки, '
               '\nбудут вытащены котлы которые отправляются Алексею Бекасово и в НГЧ-14 скринами в телеге',
               30, 130)
    list_info_for_work = ('label_bekasovo',)


def update_labels(event):
    combox = combobox.get_box()

    if combox.get() == action_list[0]:
        interface_for_normal(root)

    elif combox.get() == action_list[1]:
        interface_for_NG4(root)

    elif combox.get() == action_list[2]:
        interface_for_temperature(root)

    elif combox.get() == action_list[3]:
        interface_for_becasovo(root)


def lets_work():
    global filename
    global list_info_for_work

    try:
        if combobox.get_box().get() == action_list[0]:
            answer_file_name = by_normal(filename)
            return rdy_file_info.set_text('rdy_file_info', f'{answer_file_name}')

        elif combobox.get_box().get() == action_list[1]:
            answer_file_name = by_ng4(filename)
            return rdy_file_info.set_text('rdy_file_info', f'{answer_file_name} ')

        elif combobox.get_box().get() == action_list[2]:
            dict_entry = cl.MyEntry.entry_list
            input_min_temp_pom = int(dict_entry['input_min_temp_pom'].get())
            input_max_temp_pom = int(dict_entry['input_max_temp_pom'].get())
            input_min_temp_cont = int(dict_entry['input_min_temp_cont'].get())
            input_max_temp_cont = int(dict_entry['input_max_temp_cont'].get())
            answer_file_name = by_temperature(filename,
                                              input_min_temp_pom,
                                              input_max_temp_pom,
                                              input_min_temp_cont,
                                              input_max_temp_cont)
            temp_dict = {'max_t_pom': input_max_temp_pom,
                         'min_t_pom': input_min_temp_pom,
                         'max_t_cont': input_max_temp_cont,
                         'min_t_cont': input_min_temp_cont}

            cl.Settings.setting_obj[0].save_settings(temp_dict)
            return rdy_file_info.set_text('rdy_file_info', f'{answer_file_name}')

        elif combobox.get_box().get() == action_list[3]:
            answer_file_name = by_bek_and_ng4_14(filename)
            return rdy_file_info.set_text('rdy_file_info', f'{answer_file_name}')

    except Exception as err:
        print(err)
        rdy_file_info.set_text('rdy_file_info', 'Необходимо выбрать файл(Важно, в названии папок не должно быть скобочек())')


if __name__ == '__main__':
    root = tk.Tk()
    root.title('Сплав В.')
    root.geometry('700x300')
    root.resizable(False, False)

    label_select_file = cl.MyLabel(root, 'label_select_file', '', 130, 50)
    cl.MyButton(root, "Выбрать файл", open_file_name, 25, 50)

    combobox = cl.Combobox(root, 'cmbx', action_list, 25, 85)
    combobox.bind("<<ComboboxSelected>>", update_labels)

    cl.MyButton(root, 'Обработать/Сохранить', lets_work, 25, 250)
    rdy_file_info = cl.MyLabel(root, 'rdy_file_info', '', 175, 250)

    root.mainloop()

